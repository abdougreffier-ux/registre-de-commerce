import csv
import io
from django.utils import timezone
from django.db import transaction
from rest_framework import generics, filters, serializers, status as http_status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from django_filters.rest_framework import DjangoFilterBackend

from .models import ImmatriculationHistorique
from .serializers import (
    ImmatriculationHistoriqueListSerializer,
    ImmatriculationHistoriqueDetailSerializer,
)
from apps.core.permissions import EstAgentTribunalOuGreffier, EstGreffier, filtrer_par_auteur, est_greffier


def _next_numero_ih():
    from apps.demandes.views import _next_numero
    return _next_numero('IH')


def _validate_numero_ra_unique(numero_ra, exclude_id=None):
    """Vérifie que le N° analytique n'existe ni dans RA ni dans d'autres demandes historiques."""
    from apps.registres.models import RegistreAnalytique
    if RegistreAnalytique.objects.filter(numero_ra=numero_ra).exists():
        return False, f"Le N° analytique {numero_ra} existe déjà dans le registre analytique."
    qs = ImmatriculationHistorique.objects.filter(numero_ra=numero_ra)
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    if qs.exists():
        return False, f"Le N° analytique {numero_ra} est déjà utilisé dans une autre demande historique."
    return True, ''


def _validate_chrono_unique(numero_chrono, annee_chrono, exclude_id=None):
    """Vérifie l'unicité du couple (annee_chrono, numero_chrono)."""
    qs = ImmatriculationHistorique.objects.filter(
        numero_chrono=numero_chrono, annee_chrono=annee_chrono
    )
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    if qs.exists():
        return False, f"Le couple ({annee_chrono}/{numero_chrono}) existe déjà."
    return True, ''


# ── CRUD ──────────────────────────────────────────────────────────────────────

class HistoriqueListCreate(generics.ListCreateAPIView):
    """CDC §3.2 : immatriculation historique — agents tribunal + greffier, cloisonnement par created_by."""
    permission_classes = [EstAgentTribunalOuGreffier]
    filter_backends  = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['statut', 'type_entite', 'import_batch']
    search_fields    = ['numero_demande', 'numero_ra']
    ordering         = ['-created_at']

    def get_queryset(self):
        qs = ImmatriculationHistorique.objects.select_related(
            'localite', 'created_by', 'validated_by', 'ra'
        ).all()
        return filtrer_par_auteur(qs, self.request.user)

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return ImmatriculationHistoriqueDetailSerializer
        return ImmatriculationHistoriqueListSerializer

    def perform_create(self, serializer):
        numero_ra     = self.request.data.get('numero_ra', '').strip()
        numero_chrono = self.request.data.get('numero_chrono')
        annee_chrono  = self.request.data.get('annee_chrono')

        ok, msg = _validate_numero_ra_unique(numero_ra)
        if not ok:
            raise serializers.ValidationError({'numero_ra': msg})
        ok, msg = _validate_chrono_unique(numero_chrono, annee_chrono)
        if not ok:
            raise serializers.ValidationError({'numero_chrono': msg})

        # numero_chrono est un SerializerMethodField (lecture seule) dans le serializer,
        # il faut le passer explicitement à save() sinon il reste NULL.
        try:
            nc = int(numero_chrono) if numero_chrono not in (None, '') else None
            ac = int(annee_chrono)  if annee_chrono  not in (None, '') else None
        except (ValueError, TypeError):
            raise serializers.ValidationError(
                {'numero_chrono': 'numero_chrono et annee_chrono doivent être des entiers.'}
            )

        if nc is None:
            raise serializers.ValidationError({'numero_chrono': 'Le numéro chronologique est obligatoire.'})
        if ac is None:
            raise serializers.ValidationError({'annee_chrono': "L'année chronologique est obligatoire."})

        serializer.save(
            numero_demande=_next_numero_ih(),
            numero_ra=numero_ra,
            numero_chrono=nc,
            annee_chrono=ac,
            created_by=self.request.user,
        )


class HistoriqueDetail(generics.RetrieveUpdateAPIView):
    """CDC §3.2 : agents voient uniquement leurs dossiers."""
    permission_classes = [EstAgentTribunalOuGreffier]
    serializer_class = ImmatriculationHistoriqueDetailSerializer

    def get_queryset(self):
        qs = ImmatriculationHistorique.objects.select_related(
            'localite', 'created_by', 'validated_by', 'ra'
        ).all()
        return filtrer_par_auteur(qs, self.request.user)


# ── Actions workflow ───────────────────────────────────────────────────────────

class HistoriqueActionView(APIView):
    """CDC §6 : workflow historique.
    Actions agents : soumettre. Actions greffier : retourner, valider, annuler."""
    permission_classes = [EstAgentTribunalOuGreffier]

    ACTIONS_GREFFIER = {'retourner', 'valider', 'annuler'}

    def patch(self, request, pk, action):
        if action in self.ACTIONS_GREFFIER:
            if not EstGreffier().has_permission(request, self):
                return Response({'detail': 'Action réservée au greffier.'}, status=403)
        obj = generics.get_object_or_404(ImmatriculationHistorique, pk=pk)

        if action == 'soumettre':
            if obj.statut not in ('BROUILLON', 'RETOURNE'):
                return Response({'detail': 'Seul un brouillon ou un dossier retourné peut être soumis.'},
                                status=http_status.HTTP_400_BAD_REQUEST)
            obj.statut = 'EN_INSTANCE'
            obj.save(update_fields=['statut', 'updated_at'])
            return Response({'statut': obj.statut, 'message': 'Demande soumise au greffier.'})

        elif action == 'retourner':
            if obj.statut != 'EN_INSTANCE':
                return Response({'detail': 'Seul un dossier en instance peut être retourné.'},
                                status=http_status.HTTP_400_BAD_REQUEST)
            obs = request.data.get('observations', '').strip()
            obj.statut = 'RETOURNE'
            if obs:
                obj.observations = obs
            obj.save(update_fields=['statut', 'observations', 'updated_at'])
            return Response({'statut': obj.statut, 'message': "Dossier retourné à l'agent."})

        elif action == 'valider':
            if obj.statut != 'EN_INSTANCE':
                return Response({'detail': 'Seul un dossier en instance peut être validé.'},
                                status=http_status.HTTP_400_BAD_REQUEST)

            # Vérifier unicité N° analytique
            ok, msg = _validate_numero_ra_unique(obj.numero_ra, exclude_id=obj.id)
            if not ok:
                return Response({'detail': msg}, status=http_status.HTTP_400_BAD_REQUEST)

            # Vérifier unicité couple chrono/année
            ok, msg = _validate_chrono_unique(obj.numero_chrono, obj.annee_chrono, exclude_id=obj.id)
            if not ok:
                return Response({'detail': msg}, status=http_status.HTTP_400_BAD_REQUEST)

            obs = request.data.get('observations', '').strip()
            obj.statut       = 'VALIDE'
            obj.validated_at = timezone.now()
            obj.validated_by = request.user
            if obs:
                obj.observations = obs
            obj.save(update_fields=['statut', 'validated_at', 'validated_by', 'observations', 'updated_at'])

            # Créer les entités + RA
            try:
                with transaction.atomic():
                    ra = obj.appliquer()
            except Exception as e:
                obj.statut = 'EN_INSTANCE'
                obj.save(update_fields=['statut'])
                return Response({'detail': f"Erreur lors de la création du dossier : {e}"},
                                status=http_status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Journal d'audit
            from apps.registres.models import ActionHistorique
            ActionHistorique.objects.create(
                ra=ra,
                action='IMMATRICULATION_HISTORIQUE',
                reference_operation=obj.numero_demande,
                etat_avant={},
                etat_apres={
                    'numero_ra':           obj.numero_ra,
                    'numero_chrono':       obj.numero_chrono,
                    'annee_chrono':        obj.annee_chrono,
                    'date_immatriculation': str(obj.date_immatriculation),
                    'type_entite':         obj.type_entite,
                },
                commentaire=(
                    f"Immatriculation historique {obj.numero_demande} validée. "
                    f"RA {obj.numero_ra} créé."
                ),
                created_by=request.user,
            )
            return Response({
                'statut':    obj.statut,
                'ra_id':     ra.id,
                'ra_numero': ra.numero_ra,
                'message':   f"Dossier {ra.numero_ra} créé dans le registre analytique.",
            })

        elif action == 'rejeter':
            if obj.statut != 'EN_INSTANCE':
                return Response({'detail': 'Seul un dossier en instance peut être rejeté.'},
                                status=http_status.HTTP_400_BAD_REQUEST)
            obs = request.data.get('observations', '').strip()
            if not obs:
                return Response({'detail': 'Un motif de rejet est obligatoire.'},
                                status=http_status.HTTP_400_BAD_REQUEST)
            obj.statut = 'REJETE'
            obj.observations = obs
            obj.save(update_fields=['statut', 'observations', 'updated_at'])
            return Response({'statut': obj.statut, 'message': 'Demande rejetée.'})

        elif action == 'annuler':
            if obj.statut not in ('BROUILLON', 'RETOURNE'):
                return Response({'detail': 'Seul un brouillon ou dossier retourné peut être annulé.'},
                                status=http_status.HTTP_400_BAD_REQUEST)
            obj.statut = 'ANNULE'
            obj.save(update_fields=['statut', 'updated_at'])
            return Response({'statut': obj.statut, 'message': 'Demande annulée.'})

        return Response({'detail': 'Action inconnue.'}, status=http_status.HTTP_400_BAD_REQUEST)


# ── Import en masse ────────────────────────────────────────────────────────────

class ImportHistoriqueView(APIView):
    """
    Importe des immatriculations historiques depuis un fichier CSV ou Excel.
    POST multipart/form-data avec champ `fichier`.
    Retourne un rapport détaillé ligne par ligne.
    Réservé au greffier (CDC §3.3 — accès complet).
    """
    permission_classes = [EstGreffier]
    parser_classes = [MultiPartParser, FormParser]

    COLONNES_REQUISES = [
        'type_entite', 'numero_ra', 'numero_chrono', 'annee_chrono', 'date_immatriculation',
    ]

    def post(self, request):
        fichier = request.FILES.get('fichier')
        if not fichier:
            return Response({'detail': 'Fichier requis (champ: fichier).'},
                            status=http_status.HTTP_400_BAD_REQUEST)

        nom   = fichier.name.lower()
        rows  = []
        errors_parse = []

        if nom.endswith('.csv'):
            try:
                content = fichier.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(content))
                rows    = list(reader)
            except Exception as e:
                return Response({'detail': f'Erreur lecture CSV : {e}'},
                                status=http_status.HTTP_400_BAD_REQUEST)

        elif nom.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(fichier, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append({headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row)})
            except ImportError:
                return Response({'detail': 'openpyxl requis pour les fichiers Excel. Utilisez un CSV.'},
                                status=http_status.HTTP_400_BAD_REQUEST)
            except Exception as e:
                return Response({'detail': f'Erreur lecture Excel : {e}'},
                                status=http_status.HTTP_400_BAD_REQUEST)
        else:
            return Response({'detail': 'Format non supporté. Utilisez .csv ou .xlsx.'},
                            status=http_status.HTTP_400_BAD_REQUEST)

        if not rows:
            return Response({'detail': 'Le fichier est vide.'}, status=http_status.HTTP_400_BAD_REQUEST)

        import_batch = f"IMP-{timezone.now().strftime('%Y%m%d%H%M%S')}"
        rapport = []
        created = 0
        skipped = 0

        for i, row in enumerate(rows, start=2):
            ligne_rapport = {'ligne': i, 'numero_ra': row.get('numero_ra', ''), 'statut': 'OK', 'erreurs': []}

            # Vérifier colonnes obligatoires
            manquants = [c for c in self.COLONNES_REQUISES if not row.get(c, '').strip()]
            if manquants:
                ligne_rapport['statut']  = 'ERREUR'
                ligne_rapport['erreurs'] = [f"Colonne(s) obligatoire(s) manquante(s) : {', '.join(manquants)}"]
                rapport.append(ligne_rapport)
                skipped += 1
                continue

            numero_ra     = row['numero_ra'].strip()
            type_entite   = row['type_entite'].strip().upper()
            try:
                numero_chrono = int(row['numero_chrono'])
                annee_chrono  = int(row['annee_chrono'])
            except ValueError:
                ligne_rapport['statut']  = 'ERREUR'
                ligne_rapport['erreurs'] = ['numero_chrono et annee_chrono doivent être des entiers.']
                rapport.append(ligne_rapport)
                skipped += 1
                continue

            if type_entite not in ('PH', 'PM', 'SC'):
                ligne_rapport['statut']  = 'ERREUR'
                ligne_rapport['erreurs'] = [f"type_entite invalide : {type_entite} (valeurs : PH, PM, SC)"]
                rapport.append(ligne_rapport)
                skipped += 1
                continue

            # Vérifier doublons
            ok_ra, msg_ra = _validate_numero_ra_unique(numero_ra)
            if not ok_ra:
                ligne_rapport['statut']  = 'DOUBLON'
                ligne_rapport['erreurs'] = [msg_ra]
                rapport.append(ligne_rapport)
                skipped += 1
                continue

            ok_ch, msg_ch = _validate_chrono_unique(numero_chrono, annee_chrono)
            if not ok_ch:
                ligne_rapport['statut']  = 'DOUBLON'
                ligne_rapport['erreurs'] = [msg_ch]
                rapport.append(ligne_rapport)
                skipped += 1
                continue

            # Construire donnees
            donnees = {}
            if type_entite == 'PH':
                donnees = {
                    'nom':           row.get('nom', ''),
                    'prenom':        row.get('prenom', ''),
                    'nom_ar':        row.get('nom_ar', ''),
                    'prenom_ar':     row.get('prenom_ar', ''),
                    'nni':           row.get('nni', ''),
                    'num_passeport': row.get('num_passeport', ''),
                    'adresse':       row.get('adresse', ''),
                    'ville':         row.get('ville', ''),
                    'telephone':     row.get('telephone', ''),
                    'email':         row.get('email', ''),
                    'profession':    row.get('profession', ''),
                    'domaines':      [],
                }
            elif type_entite == 'PM':
                donnees = {
                    'denomination':    row.get('denomination', ''),
                    'denomination_ar': row.get('denomination_ar', ''),
                    'sigle':           row.get('sigle', ''),
                    'siege_social':    row.get('siege_social', ''),
                    'ville':           row.get('ville', ''),
                    'telephone':       row.get('telephone', ''),
                    'email':           row.get('email', ''),
                    'capital_social':  row.get('capital_social', ''),
                    'devise_capital':  row.get('devise_capital', 'MRU'),
                    'domaines': [], 'associes': [], 'gerants': [],
                }
            elif type_entite == 'SC':
                donnees = {
                    'denomination':    row.get('denomination', ''),
                    'denomination_ar': row.get('denomination_ar', ''),
                    'pays_origine':    row.get('pays_origine', ''),
                    'capital_affecte': row.get('capital_affecte', ''),
                    'siege_social':    row.get('siege_social', ''),
                    'ville':           row.get('ville', ''),
                    'telephone':       row.get('telephone', ''),
                    'email':           row.get('email', ''),
                    'domaines':        [],
                }

            # Créer la demande
            try:
                ImmatriculationHistorique.objects.create(
                    numero_demande=_next_numero_ih(),
                    type_entite=type_entite,
                    numero_ra=numero_ra,
                    numero_chrono=numero_chrono,
                    annee_chrono=annee_chrono,
                    date_immatriculation=row['date_immatriculation'],
                    donnees=donnees,
                    import_batch=import_batch,
                    import_row=i,
                    created_by=request.user,
                )
                created += 1
                ligne_rapport['statut'] = 'CREE'
            except Exception as e:
                ligne_rapport['statut']  = 'ERREUR'
                ligne_rapport['erreurs'] = [str(e)]
                skipped += 1

            rapport.append(ligne_rapport)

        return Response({
            'import_batch': import_batch,
            'total':   len(rows),
            'created': created,
            'skipped': skipped,
            'rapport': rapport,
        })
